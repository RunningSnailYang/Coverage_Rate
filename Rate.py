#-*- coding:utf-8 -*-
import cv2
import math
import numpy as np
import openpyxl as opxl

# This function is used to calculate the distance between two points which are
# represented by latitude and longitude.
def DistanceLatLngToDistance(Point0, Point1):
    EarthR = 6371790
    LatDistance = abs(Point0[0] - Point1[0]) * math.pi / 180 * EarthR
    AverageLat = (Point0[0] + Point1[0]) / 2 * math.pi / 180
    LngR = EarthR * math.cos(AverageLat)
    LngDistance = abs(Point0[1] - Point1[1]) * math.pi / 180 * LngR
    Distance = math.sqrt(LatDistance ** 2 + LngDistance ** 2)
    return Distance
    
Filename = 'RBook.xlsx'
ReadBook = opxl.load_workbook(Filename) # Load existing excel workbook.
ResultsBook = opxl.Workbook() # Create a new excel workbook.
ResultsSheet = ResultsBook.active # Get the active sheet.
# The method get_sheet_names() will return the names of all sheet
SheetNames = ReadBook.get_sheet_names()[1:]
CenterOverType = 'A5'
StatOverTypes = ['A1', 'A2', 'A3', 'A4', 'A6', 'B1', 'B2', 'C1', 'C2', 'C3',
                 'C4', 'D1', 'D2', 'D3', 'D4', 'D5', 'E1', 'E2', 'E3']
for OverTypeIdx, OverType in enumerate(StatOverTypes):
    # Locate a cell by the idxes of row and column.
    ResultsSheet.cell(row = 1, column = OverTypeIdx + 2).value = OverType

# SheetNames = SheetNames[0:1]
for SheetIdx, SheetName in enumerate(SheetNames):
    # print SheetName
    ResultsSheet.cell(row = SheetIdx + 2, column = 1).value = SheetName
    # The method get_sheet_by_name() is used to get sheet by the name of the sheet.
    Sheet = ReadBook.get_sheet_by_name(SheetName)
    TYPEs = [Cell.value for Cell in Sheet['F'][1:]]
    LATs = np.array([Cell.value for i, Cell in enumerate(Sheet['C'][1:]) if TYPEs[i]])
    LNGs = np.array([Cell.value for i, Cell in enumerate(Sheet['D'][1:]) if TYPEs[i]])
    SecInds = np.array([Cell.value for i, Cell in enumerate(Sheet['G'][1:]) if TYPEs[i]])
    TYPEs = np.array([TYPE for TYPE in TYPEs if TYPE])
    OverTypes = np.array([str(TYPE) + str(SecInd) for TYPE, SecInd in zip(TYPEs, SecInds)])
    KeyPointsLat = [LATs.max(), LATs.min(), (LATs.max() + LATs.min()) / 2]
    KeyPointsLng = [LNGs.max(), LNGs.min(), (LNGs.max() + LNGs.min()) / 2]
    
    Width = int(math.ceil(DistanceLatLngToDistance((KeyPointsLat[2], KeyPointsLng[0]),
                                                   (KeyPointsLat[2], KeyPointsLng[1]))))
    Height = int(math.ceil(DistanceLatLngToDistance((KeyPointsLat[0], KeyPointsLng[2]),
                                                    (KeyPointsLat[1], KeyPointsLng[2]))))
    # print Height, Width
    PointsImg = np.zeros((Height + 1, Width + 1), np.uint8)
    CircleMask = np.zeros((Height + 1, Width + 1), np.uint8)
    # np.where() will return a tuple of which each element is a ndarry and the ndarry only represents
    # the coordinates of one dim.
    CenterToIdxes = np.where(OverTypes == CenterOverType)[0]
    if not len(CenterToIdxes):
        raise NameError
    OverTypeToIdxes = dict([(OverType, np.where(OverTypes == OverType)[0]) for OverType in StatOverTypes])
    OverTypeToRate = dict([(OverType, 0) for OverType in StatOverTypes])
    # Iterate each type
    for OverTypeIdx, OverType in enumerate(StatOverTypes):
        # print 'dsb', OverType
        Idxes = OverTypeToIdxes[OverType]
        if len(Idxes):
            Rate = 0.
            S = 0
            for Idx in Idxes:
                PointY = int((LATs[Idx] - KeyPointsLat[1]) / (KeyPointsLat[0] - KeyPointsLat[1]) * Height)
                PointX = int((LNGs[Idx] - KeyPointsLng[1]) / (KeyPointsLng[0] - KeyPointsLng[1]) * Width)
                PointsImg[PointY, PointX] += 1
            # Iterate each residential area.
            for CenterIdx in CenterToIdxes:
                PointY = int((LATs[CenterIdx] - KeyPointsLat[1]) / (KeyPointsLat[0] - KeyPointsLat[1]) * Height)
                PointX = int((LNGs[CenterIdx] - KeyPointsLng[1]) / (KeyPointsLng[0] - KeyPointsLng[1]) * Width)
                # print PointY, PointX
                CircleMask = cv2.circle(CircleMask, (PointX, PointY), 250, 1, -1)
                # print 'sb', (CircleMask * PointsImg).any()
                Rate = (Rate * S + (CircleMask * PointsImg).any()) / (S + 1)
                S += 1
                CircleMask = np.zeros_like(CircleMask)
        else:
            Rate = 0.
        OverTypeToRate[OverType] = Rate
        ResultsSheet.cell(row = SheetIdx + 2, column = OverTypeIdx + 2).value = Rate
    # print OverTypeToRate
    # print len(OverTypeToRate)
ResultsBook.save('Results.xlsx')

